import customtkinter as ctk
import pdfplumber
from tkinter import filedialog
import re
from openpyxl import Workbook

class FracMasterApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("FracMaster Toolbox - Dev")
        self.geometry("1300x800")

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.tabs = ctk.CTkTabview(self)
        self.tabs.pack(expand=True, fill="both")

        self.create_job_setup_tab()
        self.create_file_dropper_tab()
        self.create_perf_converter_tab()

    def create_job_setup_tab(self):
        tab = self.tabs.add("Job Setup")
        placeholder = ctk.CTkLabel(tab, text="[Job Setup Tool Placeholder]\n(To be reintegrated later)", font=("Arial", 16))
        placeholder.pack(pady=20)

    def create_file_dropper_tab(self):
        tab = self.tabs.add("File Dropper")
        placeholder = ctk.CTkLabel(tab, text="[File Dropper Tool Placeholder]\n(To be reintegrated later)", font=("Arial", 16))
        placeholder.pack(pady=20)

    def create_perf_converter_tab(self):
        self.perf_well_data = []
        self.perf_data = {}
        self.current_well_index = 0

        tab = self.tabs.add("Perf Converter")

        self.instructions = ctk.CTkLabel(
            tab,
            text=(
                "Upload a Completion Procedure PDF. The tool will extract:\n\n"
                "- Top/Bottom Perf Depths\n"
                "- Plug Depths\n\n"
                "It now auto‐finds the “Plug” and “Cluster” columns in your header,\n"
                "so “90-270” or “Shot/Ft” no longer pollute your data."
            ),
            justify="left"
        )
        self.instructions.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=10)

        self.num_wells_entry = ctk.CTkEntry(tab, placeholder_text="How many wells on this pad?")
        self.num_wells_entry.place(x=10, y=10)
        self.load_wells_button = ctk.CTkButton(tab, text="Submit Well Count", command=self.load_well_entries)
        self.load_wells_button.place(x=10, y=50)

        self.well_entry_frame = ctk.CTkScrollableFrame(tab, height=200, width=1200)
        self.well_entry_frame.place(x=10, y=100)

        self.upload_button = ctk.CTkButton(tab, text="Upload Completion Procedure PDF", command=self.upload_pdf)
        self.upload_button.place(x=10, y=310)
        self.next_button = ctk.CTkButton(tab, text="Proceed to Next Well's Perf Data", command=self.show_next_well, state="disabled")
        self.next_button.place(x=10, y=350)

        scrollable_frame = ctk.CTkScrollableFrame(tab, height=3000, width=925)
        scrollable_frame.place(x=10, y=390)
        self.result_box = ctk.CTkTextbox(scrollable_frame, wrap="none", height=475, width=925)
        self.result_box.pack(pady=10, fill="y", expand=True)

    def load_well_entries(self):
        for w in self.well_entry_frame.winfo_children():
            w.destroy()
        try:
            count = int(self.num_wells_entry.get())
        except ValueError:
            return
        self.perf_well_data = []
        for i in range(count):
            n = ctk.CTkEntry(self.well_entry_frame, placeholder_text=f"Well {i+1} Name")
            p = ctk.CTkEntry(self.well_entry_frame, placeholder_text="Pg. # e.g. 22-25")
            s = ctk.CTkEntry(self.well_entry_frame, placeholder_text="# Stages")
            c = ctk.CTkEntry(self.well_entry_frame, placeholder_text="# Clusters/Stage")
            n.grid(row=i, column=0, padx=5, pady=2, sticky="w")
            p.grid(row=i, column=1, padx=5, pady=2, sticky="w")
            s.grid(row=i, column=2, padx=5, pady=2, sticky="w")
            c.grid(row=i, column=3, padx=5, pady=2, sticky="w")
            self.perf_well_data.append((n, p, s, c))

    def upload_pdf(self):
        pdf_path = filedialog.askopenfilename(filetypes=[("PDF Files","*.pdf")])
        if not pdf_path:
            return

        self.perf_data.clear()
        self.result_box.delete("1.0","end")
        self.current_well_index = 0

        # build mapping of well → pages + cluster count
        well_map = {}
        for name_e, pages_e, stages_e, clusters_e in self.perf_well_data:
            name = name_e.get().strip()
            if not name:
                continue
            raw = pages_e.get().replace(" ","")
            pages = set()
            for part in raw.split(';'):
                if '-' in part:
                    a,b = part.split('-',1)
                    if a.isdigit() and b.isdigit():
                        pages.update(range(int(a),int(b)+1))
                elif part.isdigit():
                    pages.add(int(part))
            try:
                ncl = int(clusters_e.get())
            except:
                ncl = 0
            well_map[name] = {'pages': sorted(pages), 'n_clusters': ncl}

        try:
            with pdfplumber.open(pdf_path) as pdf:
                for well, cfg in well_map.items():
                    pages = cfg['pages']
                    ncl   = cfg['n_clusters']
                    stages = []

                    for pnum in pages:
                        page = pdf.pages[pnum-1]
                        tables = page.extract_tables() or []

                        # 1) find a valid table
                        for tbl in tables:
                            hdr = tbl[0]
                            if not hdr:
                                continue
                            low = [ (h or "").lower() for h in hdr ]
                            if 'stage' in " ".join(low) and 'plug' in " ".join(low):
                                self.result_box.insert("end", f"🔎 Found candidate table on page {pnum}\n")
                                try:
                                    i_stage   = next(i for i,h in enumerate(low) if h.strip().startswith('stage'))
                                    i_plug    = next(i for i,h in enumerate(low) if 'plug' in h)
                                    cluster_is = [i for i,h in enumerate(low) if 'cluster' in h][:ncl]
                                except StopIteration:
                                    continue

                                for row in tbl[1:]:
                                    if not row or not row[i_stage]:
                                        continue
                                    stg_txt = row[i_stage].strip()
                                    if not stg_txt.isdigit():
                                        continue
                                    stg = f"{int(stg_txt):02d}"

                                    plug_txt = (row[i_plug] or "").replace(',','').strip()
                                    if not re.fullmatch(r"\d+(?:\.\d+)?", plug_txt):
                                        continue

                                    cl_vals = []
                                    for ci in cluster_is:
                                        ctxt = (row[ci] or "").replace(',','').strip()
                                        if re.fullmatch(r"\d+(?:\.\d+)?", ctxt):
                                            cl_vals.append(float(ctxt))
                                    if len(cl_vals) < ncl:
                                        continue

                                    plug = float(plug_txt)
                                    top  = min(cl_vals)
                                    bot  = max(cl_vals)
                                    stages.append((stg, plug, top, bot))

                                break  # stop after first matching table

                    # 2) fallback: line‐based “Stage XX …”
                    if len(stages) < (1 if ncl else 0):
                        for pnum in pages:
                            text = pdf.pages[pnum-1].extract_text() or ""
                            for L in text.splitlines():
                                m = re.match(r"Stage\s+(\d+)", L, re.IGNORECASE)
                                if not m:
                                    continue
                                toks = re.findall(r"\d+(?:\.\d+)?", L.replace(',',''))
                                if len(toks) >= 2 + ncl:
                                    stg = f"{int(toks[0]):02d}"
                                    plug = float(toks[1])
                                    clv  = list(map(float, toks[2:2+ncl]))
                                    top  = min(clv)
                                    bot  = max(clv)
                                    if not any(s[0]==stg for s in stages):
                                        stages.append((stg, plug, top, bot))

                    stages.sort(key=lambda x: int(x[0]))
                    self.perf_data[well] = stages
                    self.result_box.insert("end", f"\n=== {well}: Parsed {len(stages)} stages ===\n")
                    for s in stages:
                        self.result_box.insert("end", f"  Stage {s[0]}: plug={s[1]}, top={s[2]}, bot={s[3]}\n")

                    if stages:
                        self.next_button.configure(state="normal")

        except Exception as e:
            self.result_box.insert("end", f"❌ ERROR reading PDF: {e}\n")

    def show_next_well(self):
        wells = list(self.perf_data)
        if self.current_well_index < len(wells):
            w = wells[self.current_well_index]
            self.result_box.insert("end", f"\n>>> {w} <<<\n")
            for stg, plug, top, bot in self.perf_data[w]:
                self.result_box.insert("end", f"  {stg}: plug={plug}, top={top}, bot={bot}\n")
            self.current_well_index += 1
            if self.current_well_index >= len(wells):
                self.next_button.configure(text="Export to Excel", command=self.export_excel)
        else:
            self.export_excel()

    def export_excel(self):
        default = "Perf_Converter_" + "-".join(self.perf_data.keys()) + ".xlsx"
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default, filetypes=[("Excel","*.xlsx")])
        if not path:
            return
        wb = Workbook()
        for well, stgs in self.perf_data.items():
            ws = wb.create_sheet(title=well)
            ws.append(["Stage","Plug Depth","Top Perf","Bottom Perf"])
            for stg, plug, top, bot in stgs:
                # write numbers as real Excel numbers
                ws.append([stg, plug, top, bot])
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(path)
        self.result_box.insert("end", f"\n✅ Saved Excel: {path}\n")

if __name__ == "__main__":
    app = FracMasterApp()
    app.mainloop()