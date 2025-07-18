import customtkinter as ctk
import os
import pdfplumber
from tkinter import filedialog
import openpyxl
from openpyxl import Workbook
import re

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class FracMasterApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("FracMaster Toolbox - Dev")
        self.geometry("1300x800")

        self.tabs = ctk.CTkTabview(self)
        self.tabs.pack(expand=True, fill="both")

        self.create_job_setup_tab()
        self.create_file_dropper_tab()
        self.create_perf_converter_tab()

    def create_job_setup_tab(self):
        tab = self.tabs.add("Job Setup")
        placeholder = ctk.CTkLabel(tab, text="[Job Setup Tool Placeholder]\n(To be reintegrated later)", font=("Arial", 16))
        placeholder.pack(pady=20)

    def create_file_dropper_tab(self):
        tab = self.tabs.add("File Dropper")
        placeholder = ctk.CTkLabel(tab, text="[File Dropper Tool Placeholder]\n(To be reintegrated later)", font=("Arial", 16))
        placeholder.pack(pady=20)

    def create_perf_converter_tab(self):
        self.perf_well_data = []
        self.perf_data = {}
        self.current_well_index = 0

        tab = self.tabs.add("Perf Converter")

        self.instructions = ctk.CTkLabel(tab, text="Upload a Completion Procedure PDF. The tool will extract:\n\n- Top/Bottom Perf Depths\n- Plug Depths\n\nStage 01 always returns NULL.", justify="left")
        self.instructions.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=10)

        self.num_wells_entry = ctk.CTkEntry(tab, placeholder_text="How many wells on this pad?")
        self.num_wells_entry.place(x=10, y=10)

        self.load_wells_button = ctk.CTkButton(tab, text="Submit Well Count", command=self.load_well_entries)
        self.load_wells_button.place(x=10, y=50)

        self.well_entry_frame = ctk.CTkScrollableFrame(tab, height=200, width=1200)
        self.well_entry_frame.place(x=10, y=100)

        self.upload_button = ctk.CTkButton(tab, text="Upload Completion Procedure PDF", command=self.upload_pdf)
        self.upload_button.place(x=10, y=310)

        self.next_button = ctk.CTkButton(tab, text="Proceed to Next Well's Perf Data", command=self.show_next_well)
        self.next_button.place(x=10, y=350)
        self.next_button.configure(state="disabled")

        scrollable_frame = ctk.CTkScrollableFrame(tab, height=3000, width=925)
        scrollable_frame.place(x=10, y=390)

        self.result_box = ctk.CTkTextbox(scrollable_frame, wrap="none", height=475, width=925)
        self.result_box.pack(pady=10, fill="y", expand=True, anchor="se")

    def load_well_entries(self):
        for child in self.well_entry_frame.winfo_children():
            child.destroy()
        try:
            count = int(self.num_wells_entry.get())
            self.perf_well_data = []
            for i in range(count):
                name_entry = ctk.CTkEntry(self.well_entry_frame, placeholder_text=f"Well {i+1} Name")
                page_entry = ctk.CTkEntry(self.well_entry_frame, placeholder_text="Pg. # e.g. 22-25")
                stage_entry = ctk.CTkEntry(self.well_entry_frame, placeholder_text="# Stages")
                cluster_entry = ctk.CTkEntry(self.well_entry_frame, placeholder_text="# Clusters/Stage")

                name_entry.grid(row=i, column=0, padx=5, pady=2, sticky="w")
                page_entry.grid(row=i, column=1, padx=5, pady=2, sticky="w")
                stage_entry.grid(row=i, column=2, padx=5, pady=2, sticky="w")
                cluster_entry.grid(row=i, column=3, padx=5, pady=2, sticky="w")

                self.perf_well_data.append((name_entry, page_entry, stage_entry, cluster_entry))
        except ValueError:
            pass

    def upload_pdf(self):
        pdf_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if not pdf_path:
            return

        # Clear previous logs
        self.result_box.delete("1.0", "end")

        try:
            self.perf_data = {}
            # Build well→pages map
            well_page_map = {}
            for well_entry, pages_entry, _, _ in self.perf_well_data:
                name = well_entry.get().strip()
                pages_raw = pages_entry.get().replace(" ", "")
                page_set = set()
                for part in pages_raw.split(';'):
                    if '-' in part:
                        lo, hi = map(int, part.split('-'))
                        page_set.update(range(lo, hi+1))
                    elif part.isdigit():
                        page_set.add(int(part))
                if name:
                    well_page_map[name] = page_set

            self.result_box.insert("end", f"🔍 Well→Pages mapping: {well_page_map}\n")

            with pdfplumber.open(pdf_path) as pdf:
                for well_name, pages in well_page_map.items():
                    self.result_box.insert("end", f"\n🔍 Parsing '{well_name}' on pages {sorted(pages)}\n")
                    stages = []

                    for i, page in enumerate(pdf.pages):
                        page_num = i + 1
                        if page_num not in pages:
                            continue

                        try:
                            self.result_box.insert("end", f"\n--- Scanning page {page_num} ---\n")
                            text = page.extract_text() or ""
                            preview = text.replace("\n", " ")[:200]
                            self.result_box.insert("end", f"📝 Text preview: {preview}\n")

                            tables = page.extract_tables() or []
                            if not tables:
                                self.result_box.insert("end", "🔍 No tables found on this page\n")
                            for table in tables:
                                self.result_box.insert("end", f"🔍 Found table with {len(table)} rows\n")
                                for row in table:
                                    self.result_box.insert("end", f"    Row: {row}\n")
                                    # only data rows: first cell must be an integer stage
                                    cell0 = (row[0] or "").strip()
                                    if not cell0.isdigit():
                                        continue
                                    stage_number = cell0.zfill(2)

                                    # convert remaining cells
                                    nums = []
                                    for cell in row[1:]:
                                        if not cell:
                                            continue
                                        clean = cell.replace(",", "").strip()
                                        if re.match(r"^\d+(\.\d+)?$", clean):
                                            try:
                                                nums.append(float(clean))
                                            except ValueError:
                                                self.result_box.insert("end", f"🔍 Could not convert '{clean}' → float, skipping\n")

                                    if stage_number == "01":
                                        stages.append((stage_number, "NULL", "NULL", "NULL"))
                                    elif len(nums) >= 2:
                                        plug = str(nums[0])
                                        top  = str(min(nums[1:]))
                                        bot  = str(max(nums[1:]))
                                        stages.append((stage_number, plug, top, bot))
                                        self.result_box.insert("end", f"✅ Parsed Stage {stage_number}: Plug = {plug}, Top = {top}, Bottom = {bot}\n")

                        except Exception as e_page:
                            # log and keep going to next page
                            self.result_box.insert("end", f"❌ Error on page {page_num}: {e_page}\n")

                    self.perf_data[well_name] = stages
                    # Enable the Next button now that at least one well is parsed
                    self.next_button.configure(state="normal")

        except Exception as e:
            self.result_box.insert("1.0", f"❌ Failed to read PDF: {e}\n")

    def show_next_well(self):
        well_names = list(self.perf_data.keys())
        if self.current_well_index < len(well_names):
            well_name = well_names[self.current_well_index]
            self.result_box.insert("end", f"\n=== {well_name} Results ===\n")
            for stage, plug, top, bot in self.perf_data[well_name]:
                self.result_box.insert("end", f"  Stage {stage}: Plug = {plug}, Top = {top}, Bottom = {bot}\n")
            self.current_well_index += 1

            if self.current_well_index == len(well_names):
                self.next_button.configure(text="Export to Excel", command=self.export_excel)
        else:
            self.export_excel()

    def export_excel(self):
        well_names = list(self.perf_data.keys())
        default_name = "Perf Converter_" + "-".join(well_names) + ".xlsx"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_name
        )
        if filepath:
            wb = Workbook()
            for well in self.perf_data:
                ws = wb.create_sheet(title=well)
                ws.append(["Stage Number", "Plug Depth", "Top Perf Depth", "Bottom Perf Depth"])
                for stage, plug, top, bot in self.perf_data[well]:
                    # write numeric cells for depths
                    plug_val = None if plug == "NULL" else float(plug)
                    top_val  = None if top  == "NULL" else float(top)
                    bot_val  = None if bot  == "NULL" else float(bot)
                    ws.append([stage, plug_val, top_val, bot_val])
            # remove the default empty sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            wb.save(filepath)
            self.result_box.insert("end", f"\n✅ Excel workbook saved as: {filepath}\n")

if __name__ == "__main__":
    app = FracMasterApp()
    app.mainloop()