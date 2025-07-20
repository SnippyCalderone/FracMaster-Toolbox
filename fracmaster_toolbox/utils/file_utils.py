from openpyxl import Workbook


def save_perf_excel(perf_data: dict, path: str) -> None:
    """Save perf data mapping to an Excel workbook."""
    wb = Workbook()
    for well, stages in perf_data.items():
        ws = wb.create_sheet(title=well)
        ws.append(["Stage", "Plug Depth", "Top Perf", "Bottom Perf"])
        for stage, plug, top, bot in stages:
            ws.append([stage, plug, top, bot])
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)
